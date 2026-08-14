from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple
import re
import shutil

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .common import find_id_col, log


PROFESSOR_TEMPLATE = "professor"
CLEAN_TEMPLATE = "clean"
CUSTOM_TEMPLATE = "custom"
VALID_TEMPLATE_MODES = {PROFESSOR_TEMPLATE, CLEAN_TEMPLATE, CUSTOM_TEMPLATE}


RATING_LABEL = "Historical_Morningstar_Medalist_Rating|MMR00"
TYPE_LABEL = "Morningstar_Medalist_Rating_Type|MMR08"
RATING_DESCRIPTION = "Historical_Morningstar_Medalist_Rating"
TYPE_DESCRIPTION = "Morningstar_Medalist_Rating_Type"


def normalize_template_mode(value: str) -> str:
    mode = str(value or PROFESSOR_TEMPLATE).strip().lower()
    aliases = {
        "professor format": PROFESSOR_TEMPLATE,
        "professor": PROFESSOR_TEMPLATE,
        "clean table format": CLEAN_TEMPLATE,
        "clean": CLEAN_TEMPLATE,
        "custom workbook template": CUSTOM_TEMPLATE,
        "custom": CUSTOM_TEMPLATE,
    }
    mode = aliases.get(mode, mode)
    if mode not in VALID_TEMPLATE_MODES:
        raise RuntimeError(
            f"Unknown output template mode {value!r}. "
            f"Choose one of: {sorted(VALID_TEMPLATE_MODES)}."
        )
    return mode


def _metadata_id_column(metadata: pd.DataFrame) -> Optional[str]:
    if metadata is None or metadata.empty:
        return None
    if "_query_investment_id" in metadata.columns:
        return "_query_investment_id"
    try:
        return find_id_col(metadata)
    except Exception:
        return None


def _prepare_export_data(
    history: pd.DataFrame,
    investment_metadata: Optional[pd.DataFrame],
) -> Tuple[List[str], List[str], pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    history = history.copy() if history is not None else pd.DataFrame()
    if not history.empty:
        history["investment_id"] = history["investment_id"].astype(str).str.strip()
        history["month"] = history["month"].astype(str)

    history_ids = (
        set(history["investment_id"].dropna().astype(str).tolist())
        if not history.empty and "investment_id" in history.columns
        else set()
    )

    metadata_ids = set()
    metadata = pd.DataFrame()
    if investment_metadata is not None and not investment_metadata.empty:
        metadata = investment_metadata.copy()
        id_col = _metadata_id_column(metadata)
        if id_col:
            metadata["investment_id"] = metadata[id_col].astype(str).str.strip()
            metadata_ids = set(metadata["investment_id"].dropna().tolist())

    ids = sorted(history_ids | metadata_ids)
    months = (
        sorted(history["month"].dropna().astype(str).unique().tolist())
        if not history.empty and "month" in history.columns
        else []
    )

    rating_map = (
        history.pivot_table(
            index="investment_id", columns="month", values="rating", aggfunc="first"
        )
        if not history.empty and "rating" in history.columns
        else pd.DataFrame()
    )
    type_map = (
        history.pivot_table(
            index="investment_id", columns="month", values="rating_type", aggfunc="first"
        )
        if not history.empty and "rating_type" in history.columns
        else pd.DataFrame()
    )

    return ids, months, rating_map, type_map, metadata


def _month_end_date(month: str) -> date:
    return pd.Period(str(month), freq="M").end_time.date()


def _value_at(frame: pd.DataFrame, investment_id: str, month: str) -> object:
    if frame.empty or investment_id not in frame.index or month not in frame.columns:
        return ""
    value = frame.at[investment_id, month]
    if pd.isna(value):
        return ""
    return value


def _clear_generated_sheets(workbook: Workbook) -> None:
    generated = [
        name
        for name in workbook.sheetnames
        if re.fullmatch(r"Batch_\d+", name)
        or name in {"Generated_Summary", "First_Quant_Date", "Summary"}
    ]
    for name in generated:
        if name in workbook.sheetnames:
            del workbook[name]


def _replace_placeholders(workbook: Workbook, values: Dict[str, object]) -> None:
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                text = cell.value
                for key, value in values.items():
                    text = text.replace("{{" + key + "}}", str(value))
                cell.value = text


def _copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def _copy_template_sheet(workbook: Workbook, template_sheet: Worksheet, title: str) -> Worksheet:
    copied = workbook.copy_worksheet(template_sheet)
    copied.title = title
    max_row = max(copied.max_row, 1)
    max_col = max(copied.max_column, 1)
    for row in copied.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None
    return copied


def _style_professor_sheet(worksheet: Worksheet, months: Sequence[str]) -> None:
    worksheet.sheet_view.showGridLines = True
    worksheet.freeze_panes = "E1"
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 52
    worksheet.column_dimensions["C"].width = 3
    worksheet.column_dimensions["D"].width = 54
    for index in range(5, 5 + len(months)):
        worksheet.column_dimensions[get_column_letter(index)].width = 12

    date_fill = PatternFill("solid", fgColor="DCE6F1")
    value_fill = PatternFill("solid", fgColor="FFFFFF")
    thin_gray = Side(style="thin", color="D9E1F2")
    border = Border(bottom=thin_gray)

    for row_index in range(1, worksheet.max_row + 1):
        is_date_row = row_index % 2 == 1
        fill = date_fill if is_date_row else value_fill
        for cell in worksheet[row_index]:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        if is_date_row:
            worksheet.cell(row_index, 1).font = Font(bold=True)
            worksheet.cell(row_index, 2).font = Font(bold=True)

    for row_index in range(1, worksheet.max_row + 1, 2):
        for column_index in range(5, 5 + len(months)):
            worksheet.cell(row_index, column_index).number_format = "yyyy-mm-dd"
            worksheet.cell(row_index, column_index).alignment = Alignment(horizontal="center")


def _write_professor_batch(
    worksheet: Worksheet,
    batch_ids: Sequence[str],
    months: Sequence[str],
    rating_map: pd.DataFrame,
    type_map: pd.DataFrame,
    apply_default_style: bool = True,
) -> None:
    row_index = 1
    month_dates = [_month_end_date(month) for month in months]

    for investment_id in batch_ids:
        worksheet.cell(row_index, 1, investment_id)
        worksheet.cell(row_index, 2, RATING_LABEL)
        worksheet.cell(row_index, 4, "#NAME?")
        worksheet.cell(row_index + 1, 4, f"{investment_id} - {RATING_DESCRIPTION}")

        worksheet.cell(row_index + 2, 1, investment_id)
        worksheet.cell(row_index + 2, 2, TYPE_LABEL)
        worksheet.cell(row_index + 2, 4, "#NAME?")
        worksheet.cell(row_index + 3, 4, f"{investment_id} - {TYPE_DESCRIPTION}")

        for offset, (month, month_date) in enumerate(zip(months, month_dates), start=5):
            worksheet.cell(row_index, offset, month_date)
            worksheet.cell(row_index + 1, offset, _value_at(rating_map, investment_id, month))
            worksheet.cell(row_index + 2, offset, month_date)
            worksheet.cell(row_index + 3, offset, _value_at(type_map, investment_id, month))

        row_index += 4

    if apply_default_style:
        _style_professor_sheet(worksheet, months)
    else:
        worksheet.freeze_panes = worksheet.freeze_panes or "E1"
        for column_index in range(5, 5 + len(months)):
            for date_row in range(1, worksheet.max_row + 1, 2):
                worksheet.cell(date_row, column_index).number_format = "yyyy-mm-dd"


def _write_first_date_sheet(worksheet: Worksheet, first_by_fund: pd.DataFrame) -> None:
    if first_by_fund is None or first_by_fund.empty:
        worksheet.append(["No per-fund first-date data was generated."])
        return

    worksheet.append(list(first_by_fund.columns))
    for row in first_by_fund.itertuples(index=False, name=None):
        worksheet.append(["" if pd.isna(value) else value for value in row])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column in worksheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 40)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = width


def _write_generated_summary(
    worksheet: Worksheet,
    *,
    domicile: str,
    list_label: str,
    total_ids: int,
    funds_with_history: int,
    first_month: str,
    last_month: str,
    quantitative_month: str,
    funds_at_launch: int,
    template_mode: str,
) -> None:
    rows = [
        ("Country / domicile", domicile),
        ("List label", list_label),
        ("Total input investment IDs", total_ids),
        ("Investment IDs with quantitative history in range", funds_with_history),
        ("First month included in workbook", first_month),
        ("Last month included in workbook", last_month),
        ("First observed quantitative month", quantitative_month),
        ("Investment IDs at first quantitative month", funds_at_launch),
        ("Output template", template_mode),
        ("Qualification rule", "MMR08 = Quantitative AND MMR00 is nonblank"),
    ]
    worksheet.append(["Field", "Value"])
    for row in rows:
        worksheet.append(list(row))
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    worksheet.column_dimensions["A"].width = 48
    worksheet.column_dimensions["B"].width = 42
    worksheet.freeze_panes = "A2"


def _export_professor_or_custom(
    *,
    path: Path,
    ids: Sequence[str],
    months: Sequence[str],
    rating_map: pd.DataFrame,
    type_map: pd.DataFrame,
    first_by_fund: pd.DataFrame,
    template_mode: str,
    custom_template_file: Optional[str],
    domicile: str,
    list_label: str,
    quantitative_month: str,
    funds_at_launch: int,
) -> None:
    batch_template = None
    if template_mode == CUSTOM_TEMPLATE:
        if not custom_template_file:
            raise RuntimeError("Custom output template selected, but no Excel template was uploaded.")
        source = Path(custom_template_file).expanduser()
        if not source.exists():
            raise RuntimeError(f"Custom output template not found: {source}")
        workbook = load_workbook(source)
        batch_template = workbook["Batch_Template"] if "Batch_Template" in workbook.sheetnames else None
        _clear_generated_sheets(workbook)
        if batch_template is not None:
            batch_template.sheet_state = "hidden"
    else:
        workbook = Workbook()
        if workbook.active:
            workbook.remove(workbook.active)

    placeholders = {
        "COUNTRY": domicile,
        "DOMICILE": domicile,
        "LIST_LABEL": list_label,
        "FIRST_QUANT_MONTH": quantitative_month,
        "FUNDS_AT_LAUNCH": funds_at_launch,
        "TOTAL_INPUT_IDS": len(ids),
    }
    _replace_placeholders(workbook, placeholders)

    if template_mode == CUSTOM_TEMPLATE:
        summary_name = "Generated_Summary"
        if summary_name in workbook.sheetnames:
            del workbook[summary_name]
        summary_sheet = workbook.create_sheet(summary_name, 0)
        _write_generated_summary(
            summary_sheet,
            domicile=domicile,
            list_label=list_label,
            total_ids=len(ids),
            funds_with_history=len(rating_map.index),
            first_month=months[0] if months else "",
            last_month=months[-1] if months else "",
            quantitative_month=quantitative_month,
            funds_at_launch=funds_at_launch,
            template_mode=template_mode,
        )
        first_sheet = workbook.create_sheet("First_Quant_Date", 1)
        _write_first_date_sheet(first_sheet, first_by_fund)

    funds_per_sheet = 4000
    for batch_index, start in enumerate(range(0, len(ids), funds_per_sheet), start=1):
        title = f"Batch_{batch_index}"
        if batch_template is not None:
            worksheet = _copy_template_sheet(workbook, batch_template, title)
        else:
            worksheet = workbook.create_sheet(title)
        _write_professor_batch(
            worksheet,
            ids[start : start + funds_per_sheet],
            months,
            rating_map,
            type_map,
            apply_default_style=(batch_template is None),
        )

    if not workbook.sheetnames:
        workbook.create_sheet("Batch_1")
    workbook.save(path)


def _export_clean(
    *,
    path: Path,
    ids: Sequence[str],
    months: Sequence[str],
    rating_map: pd.DataFrame,
    type_map: pd.DataFrame,
    first_by_fund: pd.DataFrame,
    metadata: pd.DataFrame,
    domicile: str,
    list_label: str,
    quantitative_month: str,
    funds_at_launch: int,
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _write_generated_summary(
        summary,
        domicile=domicile,
        list_label=list_label,
        total_ids=len(ids),
        funds_with_history=len(rating_map.index),
        first_month=months[0] if months else "",
        last_month=months[-1] if months else "",
        quantitative_month=quantitative_month,
        funds_at_launch=funds_at_launch,
        template_mode=CLEAN_TEMPLATE,
    )

    first_sheet = workbook.create_sheet("First_Quant_Date")
    _write_first_date_sheet(first_sheet, first_by_fund)

    metadata_candidates = [
        "input_list_label",
        "input_source_file",
        "input_source_sheet",
        "input_source_row",
        "Name",
        "Investment Name",
        "Security Name",
        "Domicile",
        "query_universe",
    ]
    metadata_cols = [column for column in metadata_candidates if column in metadata.columns]
    metadata_index = (
        metadata.drop_duplicates("investment_id").set_index("investment_id")
        if not metadata.empty and "investment_id" in metadata.columns
        else pd.DataFrame()
    )

    funds_per_sheet = 5000
    for batch_index, start in enumerate(range(0, len(ids), funds_per_sheet), start=1):
        worksheet = workbook.create_sheet(f"Batch_{batch_index}")
        headers = ["Investment ID"] + metadata_cols + ["Datapoint"] + list(months)
        worksheet.append(headers)
        for investment_id in ids[start : start + funds_per_sheet]:
            base = [investment_id]
            for column in metadata_cols:
                value = (
                    metadata_index.at[investment_id, column]
                    if not metadata_index.empty and investment_id in metadata_index.index
                    else ""
                )
                base.append("" if pd.isna(value) else value)
            worksheet.append(
                base
                + [RATING_LABEL]
                + [_value_at(rating_map, investment_id, month) for month in months]
            )
            worksheet.append(
                base
                + [TYPE_LABEL]
                + [_value_at(type_map, investment_id, month) for month in months]
            )

        for cell in worksheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.column_dimensions["A"].width = 18
        for column_index in range(2, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = 18

    workbook.save(path)


def export_history_workbook(
    *,
    output_dir: Path,
    history: pd.DataFrame,
    first_by_fund: pd.DataFrame,
    list_label: str = "",
    investment_metadata: Optional[pd.DataFrame] = None,
    template_mode: str = PROFESSOR_TEMPLATE,
    custom_template_file: Optional[str] = None,
    domicile: str = "",
    quantitative_month: str = "",
    funds_at_launch: int = 0,
) -> Optional[Path]:
    """
    Export the requested full-history Excel workbook.

    professor (default): reproduces the professor workbook's four-row-per-fund
    structure and 4,000-fund Batch sheets.

    clean: creates a normal header-based table with Summary and First_Quant_Date.

    custom: preserves an uploaded workbook, replaces placeholders, and appends
    Generated_Summary, First_Quant_Date, and professor-style Batch sheets. If the
    custom workbook contains a sheet named Batch_Template, its styles/column widths
    are copied into every generated Batch sheet.
    """
    if history.empty and (investment_metadata is None or investment_metadata.empty):
        return None

    mode = normalize_template_mode(template_mode)
    ids, months, rating_map, type_map, metadata = _prepare_export_data(
        history, investment_metadata
    )
    filename_by_mode = {
        PROFESSOR_TEMPLATE: "MQR_RESULTS_PROFESSOR_FORMAT.xlsx",
        CLEAN_TEMPLATE: "MQR_RESULTS_CLEAN_TABLE.xlsx",
        CUSTOM_TEMPLATE: "MQR_RESULTS_CUSTOM_TEMPLATE.xlsx",
    }
    path = output_dir / filename_by_mode[mode]

    if mode == CLEAN_TEMPLATE:
        _export_clean(
            path=path,
            ids=ids,
            months=months,
            rating_map=rating_map,
            type_map=type_map,
            first_by_fund=first_by_fund,
            metadata=metadata,
            domicile=domicile,
            list_label=list_label,
            quantitative_month=quantitative_month,
            funds_at_launch=funds_at_launch,
        )
    else:
        _export_professor_or_custom(
            path=path,
            ids=ids,
            months=months,
            rating_map=rating_map,
            type_map=type_map,
            first_by_fund=first_by_fund,
            template_mode=mode,
            custom_template_file=custom_template_file,
            domicile=domicile,
            list_label=list_label,
            quantitative_month=quantitative_month,
            funds_at_launch=funds_at_launch,
        )

    # Keep the older filename as an alias so old scripts and saved-run views
    # continue to work, while the dashboard presents the clear filename above.
    legacy_path = output_dir / "15_medalist_history_comparable.xlsx"
    if legacy_path != path:
        shutil.copy2(path, legacy_path)

    try:
        from .cloud_storage import sync_file_to_cloud

        sync_file_to_cloud(path)
        if legacy_path.exists():
            sync_file_to_cloud(legacy_path)
    except Exception:
        pass

    log(f"Saved {path} using output template: {mode}")
    return path
