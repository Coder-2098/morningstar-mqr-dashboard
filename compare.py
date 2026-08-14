#!/usr/bin/env python3
"""
Basic Excel-to-Excel comparison for Morningstar Medalist history workbooks.

Usage:
    python3 compare.py REFERENCE.xlsx OUR_RESULT.xlsx

Optional:
    python3 compare.py REFERENCE.xlsx OUR_RESULT.xlsx \
        --month 2018-04 \
        --output SK_AFS_COMPARISON.xlsx

The script:
1. Reads all sheets whose names begin with "Batch".
2. Extracts MMR00 (rating value) and MMR08 (rating type) by fund and month.
3. Finds the first month with MMR08 = Quantitative and a nonblank MMR00.
4. Compares both workbooks at the selected launch month.
5. Prints a concise Terminal summary.
6. Creates an Excel comparison workbook.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


RecordKey = Tuple[str, str, str]  # investment_id, YYYY-MM, datapoint
Records = Dict[RecordKey, str]

DATAPOINT_CODES = ("MMR00", "MMR08")
ID_PATTERN = re.compile(r"^(?:0P|F)[A-Z0-9]{6,}$", re.IGNORECASE)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_month(value: Any) -> Optional[str]:
    """Convert Excel dates or date-like strings to YYYY-MM."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m")

    text = clean_text(value)
    if not text:
        return None

    # Exact YYYY-MM or YYYY-MM-DD.
    match = re.search(r"\b(20\d{2}|19\d{2})[-/](0[1-9]|1[0-2])(?:[-/]\d{1,2})?\b", text)
    if match:
        return f"{match.group(1)}-{match.group(2)}"

    return None


def find_datapoint(row: Iterable[Any]) -> Optional[str]:
    for value in row:
        text = clean_text(value).upper()
        for code in DATAPOINT_CODES:
            if code in text:
                return code
    return None


def find_investment_id(row: Iterable[Any]) -> Optional[str]:
    # Prefer a value that clearly looks like a Morningstar ID.
    for value in row:
        text = clean_text(value).upper()
        if ID_PATTERN.match(text):
            return text

    # Fallback: first plausible non-label value when a datapoint is present.
    ignored = {
        "INVESTMENT ID", "ID", "MSTAR FUNCTION", "DATAPOINT",
        "MMR00", "MMR08", "HISTORICAL_MORNINGSTAR_MEDALIST_RATING",
        "MORNINGSTAR_MEDALIST_RATING_TYPE",
    }
    for value in row:
        text = clean_text(value)
        upper = text.upper()
        if text and upper not in ignored and "MMR00" not in upper and "MMR08" not in upper:
            if normalize_month(text) is None and len(text) >= 6 and " " not in text:
                return upper
    return None


def nonblank(value: Any) -> bool:
    text = clean_text(value)
    return text.lower() not in {"", "nan", "none", "null", "n/a"}


def parse_batch_sheet(ws) -> Records:
    """
    Supports both common layouts:

    A. Clean two-row format:
       Header: Investment ID | Datapoint | 2018-04 | 2018-05
       Row:    F...          | ...MMR00  | Neutral | Neutral

    B. Professor-style four-row format:
       Row 1: Fund ID | ...MMR00 | date headers...
       Row 2: values...
       Row 3: Fund ID | ...MMR08 | date headers...
       Row 4: values...
    """
    rows: List[Tuple[Any, ...]] = list(ws.iter_rows(values_only=True))
    records: Records = {}

    # Find a global date header row, if one exists.
    global_date_columns: Dict[int, str] = {}
    for row in rows[:25]:
        candidate = {
            index: month
            for index, value in enumerate(row)
            if (month := normalize_month(value)) is not None
        }
        if len(candidate) >= 1:
            global_date_columns = candidate
            # A row with several dates is almost certainly the real header.
            if len(candidate) >= 2:
                break

    for row_index, row in enumerate(rows):
        datapoint = find_datapoint(row)
        if datapoint is None:
            continue

        investment_id = find_investment_id(row)
        if investment_id is None:
            continue

        # Some professor-style sheets put date headers on each datapoint row.
        local_date_columns = {
            index: month
            for index, value in enumerate(row)
            if (month := normalize_month(value)) is not None
        }
        date_columns = local_date_columns or global_date_columns
        if not date_columns:
            continue

        # In clean format, values are on the datapoint row.
        value_row = row
        same_row_values = [
            row[index] if index < len(row) else None
            for index in date_columns
        ]

        # In professor format, the datapoint row may contain dates and the
        # immediately following row contains the actual values.
        if local_date_columns and not any(
            nonblank(value) and normalize_month(value) is None
            for value in same_row_values
        ):
            if row_index + 1 < len(rows):
                value_row = rows[row_index + 1]

        for column_index, month in date_columns.items():
            value = value_row[column_index] if column_index < len(value_row) else None
            if nonblank(value) and normalize_month(value) is None:
                records[(investment_id, month, datapoint)] = clean_text(value)

    return records


def load_records(path: Path) -> Records:
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    batch_sheets = [
        name for name in workbook.sheetnames
        if name.lower().replace(" ", "_").startswith("batch")
    ]

    if not batch_sheets:
        raise ValueError(
            f"No Batch sheets found in {path.name}. "
            "Expected sheets such as Batch_1, Batch_2, etc."
        )

    records: Records = {}
    for sheet_name in batch_sheets:
        records.update(parse_batch_sheet(workbook[sheet_name]))

    if not records:
        raise ValueError(
            f"No MMR00/MMR08 monthly records could be read from {path.name}."
        )

    return records


def all_ids(records: Records) -> set[str]:
    return {investment_id for investment_id, _, _ in records}


def qualifying_ids(records: Records, month: str) -> set[str]:
    ids = all_ids(records)
    result: set[str] = set()

    for investment_id in ids:
        rating = records.get((investment_id, month, "MMR00"), "")
        rating_type = records.get((investment_id, month, "MMR08"), "")
        if nonblank(rating) and "quantitative" in rating_type.lower():
            result.add(investment_id)

    return result


def first_quantitative_month(records: Records) -> Optional[str]:
    months = sorted({month for _, month, _ in records})
    for month in months:
        if qualifying_ids(records, month):
            return month
    return None


def exact_match(left: str, right: str) -> bool:
    return clean_text(left).casefold() == clean_text(right).casefold()


def autosize(ws) -> None:
    for column_cells in ws.columns:
        max_length = max(
            (len(clean_text(cell.value)) for cell in column_cells),
            default=0,
        )
        width = min(max(max_length + 2, 12), 45)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def write_output(
    output_path: Path,
    reference_name: str,
    our_name: str,
    reference: Records,
    ours: Records,
    comparison_month: str,
) -> Dict[str, Any]:
    reference_ids = all_ids(reference)
    our_ids = all_ids(ours)
    overlap_ids = reference_ids & our_ids

    reference_quant = qualifying_ids(reference, comparison_month)
    our_quant = qualifying_ids(ours, comparison_month)
    reproduced = reference_quant & our_quant
    missing = reference_quant - our_quant
    additional = our_quant - reference_quant

    mmr00_matches = sum(
        exact_match(
            reference.get((fund_id, comparison_month, "MMR00"), ""),
            ours.get((fund_id, comparison_month, "MMR00"), ""),
        )
        for fund_id in reproduced
    )
    mmr08_matches = sum(
        exact_match(
            reference.get((fund_id, comparison_month, "MMR08"), ""),
            ours.get((fund_id, comparison_month, "MMR08"), ""),
        )
        for fund_id in reproduced
    )

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    summary_rows = [
        ("Comparison month", comparison_month),
        ("Reference workbook", reference_name),
        ("Our workbook", our_name),
        ("Reference first quantitative month", first_quantitative_month(reference) or ""),
        ("Our first quantitative month", first_quantitative_month(ours) or ""),
        ("Reference unique IDs", len(reference_ids)),
        ("Our unique IDs", len(our_ids)),
        ("IDs in both workbooks", len(overlap_ids)),
        ("Reference quantitative IDs at comparison month", len(reference_quant)),
        ("Our quantitative IDs at comparison month", len(our_quant)),
        ("Reference quantitative IDs reproduced", len(reproduced)),
        ("Reference quantitative IDs missing from ours", len(missing)),
        ("Additional quantitative IDs in ours", len(additional)),
        (
            "MMR00 exact matches on reproduced IDs",
            f"{mmr00_matches}/{len(reproduced)}"
            if reproduced else "0/0",
        ),
        (
            "MMR08 exact matches on reproduced IDs",
            f"{mmr08_matches}/{len(reproduced)}"
            if reproduced else "0/0",
        ),
    ]

    summary.append(("Metric", "Value"))
    for row in summary_rows:
        summary.append(row)
    for cell in summary[1]:
        cell.font = Font(bold=True)

    detail = workbook.create_sheet("Fund_Comparison")
    detail.append(
        (
            "Investment ID",
            "Reference MMR00",
            "Our MMR00",
            "MMR00 Match",
            "Reference MMR08",
            "Our MMR08",
            "MMR08 Match",
            "Comparison Status",
        )
    )
    for cell in detail[1]:
        cell.font = Font(bold=True)

    comparison_ids = sorted(reference_quant | our_quant)
    for fund_id in comparison_ids:
        reference_mmr00 = reference.get((fund_id, comparison_month, "MMR00"), "")
        our_mmr00 = ours.get((fund_id, comparison_month, "MMR00"), "")
        reference_mmr08 = reference.get((fund_id, comparison_month, "MMR08"), "")
        our_mmr08 = ours.get((fund_id, comparison_month, "MMR08"), "")

        if fund_id in reproduced:
            both_match = exact_match(reference_mmr00, our_mmr00) and exact_match(
                reference_mmr08, our_mmr08
            )
            status = "Exact match" if both_match else "Value differs"
        elif fund_id in missing:
            status = "Missing from our quantitative set"
        else:
            status = "Additional in our quantitative set"

        detail.append(
            (
                fund_id,
                reference_mmr00,
                our_mmr00,
                exact_match(reference_mmr00, our_mmr00)
                if fund_id in reproduced else "",
                reference_mmr08,
                our_mmr08,
                exact_match(reference_mmr08, our_mmr08)
                if fund_id in reproduced else "",
                status,
            )
        )

    distribution = workbook.create_sheet("Rating_Distribution")
    distribution.append(("Rating", "Reference Count", "Our Count"))
    for cell in distribution[1]:
        cell.font = Font(bold=True)

    reference_distribution = Counter(
        reference.get((fund_id, comparison_month, "MMR00"), "")
        for fund_id in reference_quant
    )
    our_distribution = Counter(
        ours.get((fund_id, comparison_month, "MMR00"), "")
        for fund_id in our_quant
    )
    ratings = sorted(
        (set(reference_distribution) | set(our_distribution)) - {""}
    )
    for rating in ratings:
        distribution.append(
            (rating, reference_distribution[rating], our_distribution[rating])
        )

    autosize(summary)
    autosize(detail)
    autosize(distribution)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return {
        "reference_ids": len(reference_ids),
        "our_ids": len(our_ids),
        "overlap_ids": len(overlap_ids),
        "reference_quant": len(reference_quant),
        "our_quant": len(our_quant),
        "reproduced": len(reproduced),
        "missing": len(missing),
        "additional": len(additional),
        "mmr00_matches": mmr00_matches,
        "mmr08_matches": mmr08_matches,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare two Morningstar Medalist history Excel workbooks."
    )
    parser.add_argument("reference_workbook", type=Path)
    parser.add_argument("our_workbook", type=Path)
    parser.add_argument(
        "--month",
        help="Comparison month in YYYY-MM. Default: reference launch month.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("SK_AFS_COMPARISON.xlsx"),
        help="Output Excel file. Default: SK_AFS_COMPARISON.xlsx",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    try:
        reference = load_records(args.reference_workbook)
        ours = load_records(args.our_workbook)

        reference_launch = first_quantitative_month(reference)
        our_launch = first_quantitative_month(ours)
        comparison_month = args.month or reference_launch or our_launch

        if not comparison_month:
            raise ValueError(
                "No quantitative month was found in either workbook."
            )
        if not re.fullmatch(r"\d{4}-\d{2}", comparison_month):
            raise ValueError("--month must use YYYY-MM, for example 2018-04.")

        results = write_output(
            output_path=args.output,
            reference_name=args.reference_workbook.name,
            our_name=args.our_workbook.name,
            reference=reference,
            ours=ours,
            comparison_month=comparison_month,
        )

        print()
        print("SOUTH KOREA AFS COMPARISON")
        print("=" * 64)
        print(f"Reference first quantitative month : {reference_launch or 'Not found'}")
        print(f"Our first quantitative month       : {our_launch or 'Not found'}")
        print(f"Comparison month                   : {comparison_month}")
        print(f"Reference unique IDs               : {results['reference_ids']:,}")
        print(f"Our unique IDs                     : {results['our_ids']:,}")
        print(f"IDs in both workbooks              : {results['overlap_ids']:,}")
        print(f"Reference quantitative IDs         : {results['reference_quant']:,}")
        print(f"Our quantitative IDs               : {results['our_quant']:,}")
        print(f"Reference IDs reproduced           : {results['reproduced']:,}")
        print(f"Reference IDs missing from ours    : {results['missing']:,}")
        print(f"Additional IDs in ours             : {results['additional']:,}")
        print(
            f"MMR00 exact matches                : "
            f"{results['mmr00_matches']:,}/{results['reproduced']:,}"
        )
        print(
            f"MMR08 exact matches                : "
            f"{results['mmr08_matches']:,}/{results['reproduced']:,}"
        )
        print(f"Excel comparison created           : {args.output}")
        print()
        return 0

    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
